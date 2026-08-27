import base64
import io
import re
from datetime import date, datetime

from odoo import _, api, fields, models
from odoo.exceptions import UserError

# 홈택스·스마트빌 등 **외부에서 내려받은 파일을 그대로** 올린다.
# 업체마다 헤더 문구가 달라 컬럼명을 고정할 수 없으므로, 한글 헤더를 정규화해
# 동의어 사전으로 매핑한다. 모르는 컬럼은 무시하고, 못 찾은 필수 항목만 알려준다.

# 필드 → 헤더 동의어 (정규화된 형태: 공백·괄호·특수문자 제거, 소문자)
HEADER_ALIASES = {
    "approval": ["승인번호", "국세청승인번호", "전자세금계산서승인번호", "issueid", "ntsconfirmnum"],
    "date": ["작성일자", "작성일", "발행일자", "발행일", "거래일자", "청구일자", "writedate", "issuedate"],
    "vat": ["사업자등록번호", "등록번호", "공급자등록번호", "공급받는자등록번호",
            "공급자사업자등록번호", "공급받는자사업자등록번호", "사업자번호", "corpnum"],
    "partner": ["상호", "거래처", "거래처명", "공급자상호", "공급받는자상호", "회사명", "corpname"],
    "supply": ["공급가액", "공급가", "과세표준", "supplycost"],
    "tax": ["세액", "부가세", "부가가치세", "tax"],
    "total": ["합계금액", "총액", "합계", "totalamount"],
    "item": ["품목", "품목명", "품명", "규격", "비고", "적요", "remark"],
    "origin": ["원본승인번호", "당초승인번호", "originalissueid"],
    "kind": ["종류", "세금계산서종류", "과세형태", "과세구분", "영수청구"],
}

REQUIRED = ("approval", "date", "supply")

DOC_KIND_EXEMPT = ("면세", "계산서")
DOC_KIND_ZERO = ("영세", "영세율")


def _norm(s):
    """헤더 정규화 — 공백·괄호·기호 제거 후 소문자."""
    if s is None:
        return ""
    return re.sub(r"[\s()\[\]{}·.,/\\_-]+", "", str(s)).strip().lower()


def _to_float(v):
    if v in (None, ""):
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = re.sub(r"[^\d.\-]", "", str(v))
    return float(s) if s not in ("", "-", ".") else 0.0


def _to_date(v):
    if not v:
        return False
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = re.sub(r"[^\d]", "", str(v))
    if len(s) == 8:
        try:
            return date(int(s[:4]), int(s[4:6]), int(s[6:8]))
        except ValueError:
            return False
    return False


def _clean_vat(v):
    return re.sub(r"[^\d]", "", str(v or ""))


class KrTaxInvoiceImport(models.TransientModel):
    """홈택스/스마트빌 내려받기 파일을 **그대로** 올려 세금계산서 전표를 만든다."""

    _name = "kr.tax.invoice.import"
    _description = "세금계산서 파일 반입 (홈택스·스마트빌 원본)"

    file = fields.Binary(string="내려받은 파일", required=True,
                         help="홈택스·스마트빌에서 받은 xlsx / xls / csv 를 가공 없이 그대로 올리세요.")
    filename = fields.Char()
    direction = fields.Selection(
        [("in_invoice", "매입 (우리가 받은 것)"), ("out_invoice", "매출 (우리가 발행/역발행분)")],
        string="구분", required=True, default="in_invoice")
    create_partner = fields.Boolean(
        string="미등록 거래처 자동 생성", default=False,
        help="끄면 미등록 거래처 행은 건너뛰고 목록으로 알려줍니다(권장). "
             "켜면 사업자등록번호·상호로 거래처를 새로 만듭니다.")
    post_moves = fields.Boolean(
        string="가져온 뒤 바로 게시", default=False,
        help="끄면 초안으로 두어 회계 담당이 검토 후 게시합니다(권장).")
    result = fields.Text(string="결과", readonly=True)

    # ── 파일 읽기 ──────────────────────────────────────────────
    def _rows(self):
        raw = base64.b64decode(self.file)
        name = (self.filename or "").lower()
        if name.endswith(".csv") or (not name.endswith((".xlsx", ".xls")) and b"," in raw[:200]):
            return self._rows_csv(raw)
        return self._rows_excel(raw, name)

    def _rows_csv(self, raw):
        import csv
        text = None
        for enc in ("utf-8-sig", "cp949", "euc-kr", "utf-8"):
            try:
                text = raw.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            raise UserError(_("CSV 인코딩을 인식하지 못했습니다 (UTF-8/CP949 지원)."))
        return [list(r) for r in csv.reader(io.StringIO(text))]

    def _rows_excel(self, raw, name):
        if name.endswith(".xls"):
            try:
                import xlrd
            except ImportError:
                raise UserError(_("xls 파일을 읽을 수 없습니다. xlsx 로 저장해 올려주세요."))
            book = xlrd.open_workbook(file_contents=raw)
            sh = book.sheet_by_index(0)
            return [sh.row_values(i) for i in range(sh.nrows)]
        try:
            import openpyxl
        except ImportError:
            raise UserError(_("엑셀을 읽을 수 없습니다. CSV 로 저장해 올려주세요."))
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
        return [list(r) for r in wb[wb.sheetnames[0]].iter_rows(values_only=True)]

    # ── 헤더 찾기 ─────────────────────────────────────────────
    def _find_header(self, rows):
        """머리말·안내문이 앞에 붙어 있어도 실제 헤더 행을 찾아낸다."""
        best, best_hit, best_map = None, 0, {}
        for idx, row in enumerate(rows[:30]):
            cols = {}
            for pos, cell in enumerate(row):
                n = _norm(cell)
                if not n:
                    continue
                for field, aliases in HEADER_ALIASES.items():
                    if field in cols:
                        continue
                    if any(n == a or (len(a) > 3 and a in n) for a in aliases):
                        cols[field] = pos
                        break
            hit = len(cols)
            if hit > best_hit:
                best, best_hit, best_map = idx, hit, cols
        if best is None or not all(f in best_map for f in REQUIRED):
            found = ", ".join(sorted(best_map)) or "(없음)"
            raise UserError(_(
                "헤더를 인식하지 못했습니다.\n인식된 항목: %s\n"
                "최소한 승인번호·작성일자·공급가액 컬럼이 필요합니다.\n"
                "파일 첫 행이 헤더인지 확인하시고, 계속 실패하면 이 파일을 개발에 전달해 주세요 "
                "(헤더 문구를 사전에 추가합니다).") % found)
        return best, best_map

    # ── 실행 ──────────────────────────────────────────────────
    def action_import(self):
        self.ensure_one()
        rows = self._rows()
        if not rows:
            raise UserError(_("파일이 비어 있습니다."))
        head_idx, cmap = self._find_header(rows)

        Move = self.env["account.move"]
        Partner = self.env["res.partner"]
        company = self.env.company
        is_sale = self.direction == "out_invoice"

        # 기존 승인번호 (중복 반입 차단 — DB 제약과 별개로 미리 걸러 결과를 알려준다)
        existing = set(Move.search([("kr_approval_number", "!=", False)]).mapped("kr_approval_number"))

        created, dup, skipped, errors = [], 0, [], []
        seen_in_file = set()

        for line_no, row in enumerate(rows[head_idx + 1:], start=head_idx + 2):
            def cell(field):
                pos = cmap.get(field)
                return row[pos] if pos is not None and pos < len(row) else None

            approval = str(cell("approval") or "").strip()
            if not approval:
                continue  # 빈 줄·합계 줄
            if approval in existing:
                dup += 1
                continue
            if approval in seen_in_file:
                dup += 1
                continue

            d = _to_date(cell("date"))
            if not d:
                errors.append(_("%s행: 작성일자를 읽지 못했습니다 (%s)") % (line_no, cell("date")))
                continue

            supply = _to_float(cell("supply"))
            tax_amt = _to_float(cell("tax"))
            total = _to_float(cell("total"))
            if total and abs((supply + tax_amt) - total) > 1.0:
                errors.append(_("%s행: 공급가액+세액(%s) 과 합계금액(%s) 이 다릅니다")
                              % (line_no, supply + tax_amt, total))
                continue

            vat = _clean_vat(cell("vat"))
            pname = str(cell("partner") or "").strip()
            partner = Partner.browse()
            if vat:
                partner = Partner.search(
                    [("vat", "in", (vat, self._dashed(vat)))], limit=1)
            if not partner and pname:
                partner = Partner.search([("name", "=", pname)], limit=1)
            if not partner:
                if not self.create_partner:
                    skipped.append(_("%s행: 미등록 거래처 %s (%s)")
                                   % (line_no, pname or "(상호없음)", self._dashed(vat) or "-"))
                    continue
                partner = Partner.create({
                    "name": pname or self._dashed(vat) or _("미상 거래처"),
                    "vat": self._dashed(vat) or False,
                    "company_type": "company",
                    ("supplier_rank" if not is_sale else "customer_rank"): 1,
                })

            kind = str(cell("kind") or "")
            if tax_amt > 0:
                kr_tax_type, doc_type = "taxable", "tax_invoice"
            elif any(k in kind for k in DOC_KIND_ZERO):
                kr_tax_type, doc_type = "zero", "tax_invoice"
            elif any(k in kind for k in DOC_KIND_EXEMPT):
                kr_tax_type, doc_type = "exempt", "invoice"
            else:
                kr_tax_type, doc_type = "exempt", "tax_invoice"

            tax = self._pick_tax(is_sale, tax_amt, supply, company)
            if tax_amt > 0 and not tax:
                # 세액이 있는데 세금 코드를 못 찾으면 **만들지 않는다** —
                # 그대로 넣으면 부가세가 조용히 누락된 전표가 된다
                errors.append(_("%s행: 세액 %s 에 맞는 세금 코드를 찾지 못했습니다 "
                                "(회계 설정 › 세금 확인 후 다시 시도)") % (line_no, tax_amt))
                continue
            label = str(cell("item") or "").strip() or _("세금계산서 반입")

            # 한국 세목은 기본이 **내부포함(price included)** 이다 —
            # 그 경우 단가에 세액을 포함한 금액을 넣어야 공급가액/세액이 파일과 일치한다
            price_unit = supply + tax_amt if (tax and tax.price_include) else supply
            vals = {
                "move_type": self.direction,
                "partner_id": partner.id,
                "invoice_date": d,
                "kr_approval_number": approval,
                "kr_doc_type": doc_type,
                "kr_tax_type": kr_tax_type,
                "kr_tax_type_manual": True,
                "invoice_line_ids": [(0, 0, {
                    "name": label,
                    "quantity": 1,
                    "price_unit": price_unit,
                    "tax_ids": [(6, 0, tax.ids)] if tax else [(5, 0, 0)],
                })],
            }
            origin = str(cell("origin") or "").strip()
            if origin:
                vals["kr_origin_number"] = origin
            try:
                mv = Move.create(vals)
                created.append(mv.id)
                seen_in_file.add(approval)
            except Exception as e:  # noqa: BLE001 — 한 행 실패가 전체를 막지 않게
                errors.append(_("%s행: 생성 실패 — %s") % (line_no, str(e)[:120]))

        moves = Move.browse(created)
        if self.post_moves and moves:
            try:
                moves.action_post()
            except Exception as e:  # noqa: BLE001
                errors.append(_("게시 실패(초안으로 남김) — %s") % str(e)[:200])

        self.result = self._summary(cmap, len(created), dup, skipped, errors)
        if not created:
            return {"type": "ir.actions.act_window", "res_model": self._name,
                    "res_id": self.id, "view_mode": "form", "views": [[False, "form"]],
                    "target": "new"}
        return {
            "type": "ir.actions.act_window",
            "name": _("반입된 세금계산서"),
            "res_model": "account.move",
            "view_mode": "list,form",
            "views": [[False, "list"], [False, "form"]],
            "domain": [("id", "in", created)],
        }

    # ── 도우미 ────────────────────────────────────────────────
    @staticmethod
    def _dashed(vat):
        v = _clean_vat(vat)
        return "%s-%s-%s" % (v[:3], v[3:5], v[5:]) if len(v) == 10 else (v or "")

    def _pick_tax(self, is_sale, tax_amt, supply, company):
        """세액이 있으면 10% 세금, 없으면 세금 없음. 회사 기본 세금을 우선 사용."""
        if tax_amt <= 0:
            return self.env["account.tax"].browse()
        use = "sale" if is_sale else "purchase"
        rate = round(tax_amt / supply * 100.0) if supply else 10
        Tax = self.env["account.tax"]
        base = [("type_tax_use", "=", use), ("company_id", "=", company.id),
                ("amount_type", "=", "percent")]
        cands = Tax.search(base + [("amount", "=", rate)])
        if not cands:
            cands = Tax.search(base + [("amount", "=", 10)])
        # 파일의 공급가액은 **세액 별도** 금액이므로 내부포함(price included) 세금을 쓰면
        # 1,000,000 이 909,091+90,909 로 쪼개진다 → 외부세 코드를 우선 고른다
        external = cands.filtered(lambda t: not t.price_include)
        return (external or cands)[:1]

    def _summary(self, cmap, n_created, dup, skipped, errors):
        lines = [
            _("인식된 컬럼: %s") % ", ".join(sorted(cmap)),
            _("생성 %(n)d건 / 중복(건너뜀) %(d)d건 / 거래처 미등록(건너뜀) %(s)d건 / 오류 %(e)d건")
            % {"n": n_created, "d": dup, "s": len(skipped), "e": len(errors)},
        ]
        if skipped:
            lines.append("\n" + _("[미등록 거래처 — 등록 후 다시 올리거나 '자동 생성' 옵션 사용]"))
            lines += skipped[:30]
            if len(skipped) > 30:
                lines.append(_("...외 %d건") % (len(skipped) - 30))
        if errors:
            lines.append("\n" + _("[오류]"))
            lines += errors[:30]
            if len(errors) > 30:
                lines.append(_("...외 %d건") % (len(errors) - 30))
        return "\n".join(lines)
